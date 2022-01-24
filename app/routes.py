from flask import render_template, request, Response
from app import app
from api import get_user_info, get_user_matches, get_stats

import base64
import json

from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook


def decode(token):
	try:
		decoded = base64.b64decode(token.encode('utf-8')).decode('utf-8')
		print(f"[~] Got JSON from user: {decoded}")
		return json.loads(decoded)
	except Exception as error:
		print(f"[-] ERROR: {error}")
		return None


def get_all(token):
	if token is None:
		raise Exception('No token specified.')
	decoded = decode(token)
	if decoded is None:
		raise Exception('Wrong token.')
	try:
		info = get_user_info(decoded['nickname'])
		if info is None:
			raise Exception(f"Couldn't get user info for nickname: {decoded['nickname']}.")
		user_args = {
			'user_id': info['id']
		}
		if decoded['select'] == 'last-matches':
			user_args['last'] = decoded['select-input']
		elif decoded['select'] == 'from-date':
			user_args['starting_from'] = decoded['select-input']
		elif decoded['select'] == 'in-last':
			user_args['in_last'] = decoded['select-input']
		matches, start_elo = get_user_matches(**user_args)
		stats = get_stats(matches, start_elo)
		return info, matches, start_elo, stats

	except Exception as error:
		print(f"[-] ERROR: {error}")
		raise Exception('Unknown error.')


def get_info(token):
	if token is None:
		raise Exception('No token specified.')
	decoded = decode(token)
	if decoded is None:
		raise Exception('Wrong token.')
	try:
		info = get_user_info(decoded['nickname'])
		if info is None:
			raise Exception(f"Couldn't get user info for nickname: {decoded['nickname']}.")		
		return info
	except Exception as error:
		print(f"[-] ERROR: {error}")
		raise Exception('Unknown error.')


@app.route('/')
@app.route('/index')
def index():
	return render_template('main.html')


@app.route('/widget')
def widget():
	token = request.args.get("token")
	try:
		info, matches, start_elo, stats = get_all(token)
	except Exception as error:
		return {'error': True, 'message': str(error)}
	return render_template('presets/RachelR.html', token=token, stats=stats)


@app.route('/table')
def table():
	token = request.args.get("token")
	try:
		info, matches, start_elo, stats = get_all(token)
	except Exception as error:
		return {'error': True, 'message': str(error)}
	
	print(f"[~] Creating table...")

	wb = load_workbook(filename="tables/RachelR.xlsx")
	ws = wb['Статистика']
	for row, match in enumerate(reversed(matches), start=2):
		ws.cell(column=1, row=row, value=match['kills'])
		ws.cell(column=2, row=row, value=match['deaths'])
		ws.cell(column=3, row=row, value=match['rounds'])
		ws.cell(column=4, row=row, value='win' if match['win'] else 'lose')
	ws['F2'] = start_elo
	ws['G2'] = matches[0]['elo']

	return Response(
		save_virtual_workbook(wb),
		headers={
        	'Content-Disposition': 'attachment; filename=table.xlsx',
        	'Content-type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
	)